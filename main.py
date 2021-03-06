
#
# Software for analysing the behaviour of animals, specifically 
# licking and biting from pre-recorded videos.
#
# Alison Symon for the spinal cord group at the University of Glasgow
# 2019
# 
# Based on heavily on the vlc-PyQt5-example found at:
# https://github.com/devos50/vlc-pyqt5-example
#
# This program is free software; you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation; either version 3git of the License, or
# (at your option) any later version.
#



import sys
import os.path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from PyQt5.QtCore import Qt, QTimer, QRect
from PyQt5.QtGui import QPalette, QColor, QKeySequence
from PyQt5.QtWidgets import QMainWindow, QWidget, QFrame, QSlider, QHBoxLayout, QPushButton, \
    QVBoxLayout, QAction, QFileDialog, QApplication, QLineEdit, QLabel, QShortcut, QButtonGroup
import vlc

class Player(QMainWindow):

    def __init__(self, master=None):
        QMainWindow.__init__(self, master)
        self.setWindowTitle('Spinal Group Animal Behaviour Analyser')

        # creating a basic vlc instance
        self.instance = vlc.Instance()
        # creating an empty vlc media player
        self.mediaplayer = self.instance.media_player_new()

        self.createUI()
        self.isPaused = False
        self.isLicking = False
        self.isBiting = False 

        self.startLickTime = []
        self.stopLickTime = []
        self.startBiteTime = []
        self.stopBiteTime = []
        self.includeBite = [] 
        self.includeLick = []

        self.nBites = 0
        self.nLicks = 0 
        

    def keyPressEvent(self, e):
        if e.key() == Qt.Key_Equal or e.key() == Qt.Key_Enter:
            self.resetSpeed()
        elif e.key() == Qt.Key_BracketLeft or  e.key() == Qt.Key_Plus:
            self.decreaseSpeed()
        elif e.key() == Qt.Key_BracketRight or  e.key() == Qt.Key_Minus:
            self.increaseSpeed()
        elif e.key() == Qt.Key_D:
            self.moveframeforward()
        elif e.key() == Qt.Key_A:
            self.moveframebackward()

    def createUI(self):
        # Set up the user interface, signals & slots
        self.widget = QWidget(self)
        self.setCentralWidget(self.widget)
        self.videoframe = QFrame()
        self.palette = self.videoframe.palette()
        self.palette.setColor (QPalette.Window,
                               QColor(0,0,0))
        self.videoframe.setPalette(self.palette)
        self.videoframe.setAutoFillBackground(True)

        self.videocontolbox1 = QHBoxLayout()
        self.playbutton = QPushButton('Play')
        QShortcut(QKeySequence(Qt.Key_Space), self).activated.connect(self.PlayPause)
        self.videocontolbox1.addWidget(self.playbutton)
        self.playbutton.clicked.connect(self.PlayPause)

        self.positionslider = QSlider(Qt.Horizontal, self)
        self.videocontolbox1.addWidget(self.positionslider)
        self.positionslider.setToolTip('Position')
        self.positionslider.setMaximum(1000)
        self.positionslider.sliderMoved.connect(self.setPosition)

        self.videocontolbox = QHBoxLayout()
        self.videocontolbox.addStretch(1)

        self.SpeedLabel = QLabel(self)
        self.SpeedLabel.setText(' Playback Speed: ')
        self.videocontolbox.addWidget(self.SpeedLabel)
        self.speedslider = QSlider(Qt.Horizontal, self)
        self.speedslider.setMaximum(200)
        self.speedslider.setMinimum(50)
        self.speedslider.setValue(self.mediaplayer.get_rate()*100)
        self.speedslider.setToolTip('Speed')
        self.videocontolbox.addWidget(self.speedslider)
        self.speedslider.valueChanged.connect(self.setSpeed)
        
        self.VolLabel = QLabel(self)
        self.VolLabel.setText(' Volume: ')
        self.videocontolbox.addWidget(self.VolLabel)
        self.volumeslider = QSlider(Qt.Horizontal, self)
        self.volumeslider.setMaximum(100)
        self.volumeslider.setValue(self.mediaplayer.audio_get_volume())
        self.setVolume(0)
        self.volumeslider.setToolTip('Volume')
        self.videocontolbox.addWidget(self.volumeslider)
        self.volumeslider.valueChanged.connect(self.setVolume)

        self.LickBehaviour = QHBoxLayout()
        self.licktoggle = QPushButton('Start Lick')
        self.licktoggle.setStyleSheet("color:#01544D")
        self.LickBehaviour.addWidget(self.licktoggle)
        QShortcut(QKeySequence(Qt.Key_L), self).activated.connect(self.LickStartStop)
        self.licktoggle.clicked.connect(self.LickStartStop)

        self.inicatorLickBehaviour = QHBoxLayout()

        self.LickBehaviour.addLayout(self.inicatorLickBehaviour)
        self.LickBehaviour.addStretch()
        self.lickButtonGroup = QButtonGroup()
        self.lickButtonGroup.buttonClicked.connect(self.modifydeleteLick)


        self.BiteBehaviour = QHBoxLayout()
        self.bitetoggle = QPushButton('Start Bite')
        self.bitetoggle.setStyleSheet("color:#600D07")
        QShortcut(QKeySequence(Qt.Key_B), self).activated.connect(self.BiteStartStop)
        self.BiteBehaviour.addWidget(self.bitetoggle)
        self.bitetoggle.clicked.connect(self.BiteStartStop)


        self.inicatorBiteBehaviour = QHBoxLayout()

        self.BiteBehaviour.addLayout(self.inicatorBiteBehaviour)
        self.BiteBehaviour.addStretch()
        self.biteButtonGroup = QButtonGroup()
        self.biteButtonGroup.buttonClicked.connect(self.modifydeleteBite)



        self.InfoSave = QHBoxLayout()
        self.InfoSave.addStretch(20)
        self.IDLabel = QLabel(self)
        self.IDLabel.setText('Animal ID: ')
        self.InfoSave.addWidget(self.IDLabel)
        self.animalID = QLineEdit(self)
        self.animalID.setText('')
        self.InfoSave.addWidget(self.animalID)
        self.InfoSave.addStretch(1)

        self.DateLabel = QLabel(self)
        self.DateLabel.setText('Date: ')
        self.InfoSave.addWidget(self.DateLabel)
        self.Date = QLineEdit(self)
        self.Date.setText('')
        self.InfoSave.addWidget(self.Date)

        self.InfoSave.addStretch(1)
        self.savebutton = QPushButton('Save')
        QShortcut(QKeySequence('Ctrl+S'), self).activated.connect(self.SaveData)
        self.savebutton.clicked.connect(self.SaveData)
        self.InfoSave.addWidget(self.savebutton)

        self.vboxlayout = QVBoxLayout()
        self.vboxlayout.addWidget(self.videoframe)
        self.vboxlayout.addLayout(self.videocontolbox1)
        self.vboxlayout.addLayout(self.videocontolbox)
        self.vboxlayout.addLayout(self.LickBehaviour)
        self.vboxlayout.addLayout(self.BiteBehaviour)
        self.vboxlayout.addLayout(self.InfoSave)
        self.widget.setLayout(self.vboxlayout)

        open = QAction('&Open', self)
        open.triggered.connect(self.OpenFile)
        exit = QAction('&Exit', self)
        exit.triggered.connect(sys.exit)
        menubar = self.menuBar()
        filemenu = menubar.addMenu('&File')
        filemenu.addAction(open)
        filemenu.addSeparator()
        filemenu.addAction(exit)

        self.timer = QTimer(self)
        self.timer.setInterval(200)
        self.timer.timeout.connect(self.updateUI)

    def PlayPause(self):
        # Toggle play/pause status
        if self.mediaplayer.is_playing():
            self.mediaplayer.pause()
            self.playbutton.setText('Play')
            self.isPaused = True
        else:
            if self.mediaplayer.play() == -1:
                self.OpenFile()
                return
            self.mediaplayer.play()
            self.playbutton.setText('Pause')
            self.timer.start()
            self.isPaused = False

    def Stop(self):
        # Stop player
        self.mediaplayer.stop()
        self.playbutton.setText('Play')

    def OpenFile(self, filename=None):
        # Open a media file in a MediaPlayer
        filename = QFileDialog.getOpenFileName(self, 'Open File', os.path.expanduser('~'))[0]
        if not filename:
            return

        # create the media
        if sys.version < '3':
            filename = unicode(filename)
        self.media = self.instance.media_new(filename)
        # put the media in the media player
        self.mediaplayer.set_media(self.media)
        # the media player has to be 'connected' to the QFrame
        # (otherwise a video would be displayed in it's own window)
        # you have to give the id of the QFrame (or similar object) to
        # vlc, different platforms have different functions for this
        self.mediaplayer.set_hwnd(self.videoframe.winId())
        self.PlayPause()


    def setVolume(self, Volume):
        # Set the volume
        self.mediaplayer.audio_set_volume(Volume)
        self.volumeslider.setValue(self.mediaplayer.audio_get_volume())

    def increaseSpeed(self):
        # increase the speed
        current = self.mediaplayer.get_rate()*100
        new = current + 25
        if new > 200:
            new = 200
        self.setSpeed(new)

    def decreaseSpeed(self):
        # decrease the speed
        current = self.mediaplayer.get_rate()*100
        new = current - 25
        if new < 50:
            new = 50
        self.setSpeed(new)

    def resetSpeed(self):
        self.setSpeed(100)


    def setSpeed(self, Speed):
        # Set the playback speed
        self.mediaplayer.set_rate((Speed/100))
        self.speedslider.setValue(self.mediaplayer.get_rate()*100)

    def setPosition(self, position):
        # Set the position
        # setting the position to where the slider was dragged
        self.mediaplayer.set_position(position / 1000.0)
        # the vlc MediaPlayer needs a float value between 0 and 1, Qt
        # uses integer variables, so you need a factor; the higher the
        # factor, the more precise are the results
        # (1000 should be enough)

    def LickStartStop(self):
        # Toggle state of licking behaviour
        if self.mediaplayer.is_playing():
            if self.isLicking == False:
                self.isLicking = True
                self.licktoggle.setText('Stop Lick')
                self.startLickTime.append(self.mediaplayer.get_time()/1000)
            else:
                self.isLicking = False
                self.stopLickTime.append(self.mediaplayer.get_time()/1000)
                self.includeLick.append(True)
                self.licktoggle.setText('Start Lick')

                self.lickadd = QPushButton()
                self.lickadd.setStyleSheet('background-color:#4D8782')
                self.lickadd.resize(1,1)
                self.inicatorLickBehaviour.addWidget(self.lickadd)

                self.lickButtonGroup.addButton(self.lickadd,self.nLicks)
                self.nLicks = self.nLicks + 1

    def BiteStartStop(self):
        # Toggle state of licking behaviour
        if self.mediaplayer.is_playing():
            if self.isBiting == False:
                self.isBiting = True
                self.bitetoggle.setText('Stop Bite')
                self.startBiteTime.append(self.mediaplayer.get_time()/1000)
                self.mediaplayer.get_position()
                
            else:
                self.isBiting = False
                self.stopBiteTime.append(self.mediaplayer.get_time()/1000)
                self.includeBite.append(True)
                self.bitetoggle.setText('Start Bite')


                self.biteadd = QPushButton()
                self.biteadd.setStyleSheet('background-color:#8F5551')
                self.biteadd.resize(1,1)
                self.inicatorBiteBehaviour.addWidget(self.biteadd)

                self.biteButtonGroup.addButton(self.biteadd,self.nBites)
                self.nBites = self.nBites + 1


    def modifydeleteBite(self, button):
        for item in self.biteButtonGroup.buttons():
            if button is item:
                biteNumber = self.biteButtonGroup.id(button)
                if self.includeBite[biteNumber]:
                    self.includeBite[biteNumber] = False
                    button.setStyleSheet('background-color:red')
                else:
                    self.includeBite[biteNumber] = True
                    button.setStyleSheet('background-color:#8F5551')

    def modifydeleteLick(self, button):
        for item in self.lickButtonGroup.buttons():
            if button is item:
                lickNumber = self.lickButtonGroup.id(button)
                if self.includeLick[lickNumber]:
                    self.includeLick[lickNumber] = False
                    button.setStyleSheet('background-color:red')
                else:
                    self.includeLick[lickNumber] = True
                    button.setStyleSheet('background-color:#4D8782')

                


    def moveframeforward(self):
        ct = self.mediaplayer.get_time()
        self.mediaplayer.set_time(ct+40)

    def moveframebackward(self):
        ct = self.mediaplayer.get_time()
        self.mediaplayer.set_time(ct-40)

    def SaveData(self):
        date = self.Date.text().replace('/','.')
        date = date.replace('\\','.')
        dest_filename = self.animalID.text() + '_' + date + '.xlsx'
        savefilename, _ = QFileDialog.getSaveFileName(self, 'Save File', os.path.expanduser('~') + '\\' + dest_filename, '*.xlsx')
        if savefilename:
            print(savefilename)
            wb = Workbook()
            ws1 = wb.active
            ws1.title = 'summary'
            ws2 = wb.create_sheet(title='LickData')
            ws3 = wb.create_sheet(title='BiteData')

            # Set up summary sheet
            _ = ws1.cell(column=1,row=1,value='Summary Sheet')
            _ = ws1.cell(column=1,row=2,value='Animal ID')
            _ = ws1.cell(column=2,row=2,value=self.animalID.text())
            _ = ws1.cell(column=1,row=3,value='Date')
            _ = ws1.cell(column=2,row=3,value=self.Date.text())

            if len(self.startLickTime) > 0:
                # Set up and fill in Lick Data
                _ = ws2.cell(column=1, row=1, value='Occurence')
                _ = ws2.cell(column=2, row=1, value='Start Time')
                _ = ws2.cell(column=3, row=1, value='Stop Time')
                _ = ws2.cell(column=4, row=1, value='Duration')

                xrow = 2

                for row in range(len(self.startLickTime)):
                    if self.includeLick[row]:
                        _ = ws2.cell(column=1, row=xrow, value=row)
                        _ = ws2.cell(column=2, row=xrow, value=self.startLickTime[row])
                        _ = ws2.cell(column=3, row=xrow, value=self.stopLickTime[row])
                        _ = ws2.cell(column=4, row=xrow, value='=C{0}-B{0}'.format(xrow))
                        xrow = xrow + 1
                _ = ws2.cell(column=4, row=xrow, value='=SUM(D2:D{0})'.format(xrow-1))
                
                # Create summary for Lick Data
                _ = ws1.cell(column=2,row=5,value='Lick Data')
                _ = ws1.cell(column=1,row=6,value='Occurances')
                _ = ws1.cell(column=2,row=6,value='=LickData!A{0}'.format(xrow-1))
                _ = ws1.cell(column=1,row=7,value='Total Duration')
                _ = ws1.cell(column=2,row=7,value='=LickData!D{0}'.format(xrow))
                _ = ws1.cell(column=1,row=8,value='Average Duration')
                _ = ws1.cell(column=2,row=8,value='=B7/B6')
                _ = ws1.cell(column=1,row=9,value='Run Time')
                _ = ws1.cell(column=2,row=9,value=self.media.get_duration())
                _ = ws1.cell(column=1,row=10,value='Frequency')
                _ = ws1.cell(column=2,row=10,value='=B6/B9')

            if len(self.startBiteTime) > 0:
                # Set up and fill in Bite Data
                _ = ws3.cell(column=1, row=1, value='Occurence')
                _ = ws3.cell(column=2, row=1, value='Start Time')
                _ = ws3.cell(column=3, row=1, value='Stop Time')
                _ = ws3.cell(column=4, row=1, value='Duration')

                xrow = 2

                for row in range(len(self.startBiteTime)):
                    if self.includeBite[row]:
                        _ = ws3.cell(column=1, row=xrow, value=xrow-1)
                        _ = ws3.cell(column=2, row=xrow, value=self.startBiteTime[row])
                        _ = ws3.cell(column=3, row=xrow, value=self.stopBiteTime[row])
                        _ = ws3.cell(column=4, row=xrow, value='=C{0}-B{0}'.format(xrow))
                        xrow = xrow + 1
                _ = ws3.cell(column=4, row=xrow, value='=SUM(D2:D{0})'.format(xrow-1))
                
                # Create summary for Bite Data
                _ = ws1.cell(column=3,row=5,value='Bite Data')
                _ = ws1.cell(column=3,row=6,value='=BiteData!A{0}'.format(xrow-1))
                _ = ws1.cell(column=3,row=7,value='=BiteData!D{0}'.format(xrow))
                _ = ws1.cell(column=3,row=8,value='=B7/B6')
                _ = ws1.cell(column=3,row=9,value=self.media.get_duration())
                _ = ws1.cell(column=3,row=10,value='=B6/B9')
                
            ws1.column_dimensions['A'].width = 15
            
            wb.save(savefilename)
        else:
            return

    def updateUI(self):
        # updates the user interface
        # setting the slider to the desired position
        
        self.positionslider.setValue(self.mediaplayer.get_position() * 1000)

        if not self.mediaplayer.is_playing():
            # no need to call this function if nothing is played
            self.timer.stop()
            if not self.isPaused:
                # after the video finished, the play button stills shows
                # 'Pause', not the desired behavior of a media player
                # this will fix it
                self.Stop()

if __name__ == '__main__':
    app = QApplication(sys.argv)
    player = Player()
    player.show()
    player.resize(1280, 960)
    if sys.argv[1:]:
        player.OpenFile(sys.argv[1])
    sys.exit(app.exec_())